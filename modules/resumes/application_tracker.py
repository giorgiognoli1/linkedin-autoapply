#!/usr/bin/env python3
"""
Application Tracker Module
Logs all job applications to an Excel file.
"""

import os
from datetime import datetime

try:
    from openpyxl import Workbook, load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from modules.helpers import print_lg

# Excel file path
EXCEL_FILE = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "all excels", "applications_log.xlsx")


def log_application(
    job_title: str,
    company: str,
    application_date: datetime,
    resume_path: str,
    job_link: str,
    status: str = "Applied"
) -> bool:
    """
    Log a job application to the Excel tracker.
    
    Args:
        job_title: Title of the job
        company: Company name
        application_date: When the application was submitted
        resume_path: Path to the generated resume PDF
        job_link: URL to the job posting
        status: Application status (default: "Applied")
        
    Returns:
        True if logged successfully, False otherwise
    """
    if not OPENPYXL_AVAILABLE:
        print_lg("⚠️ openpyxl not installed. Skipping Excel logging.")
        return False
    
    try:
        # Ensure directory exists
        os.makedirs(os.path.dirname(EXCEL_FILE), exist_ok=True)
        
        # Load or create workbook
        if os.path.exists(EXCEL_FILE):
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Applications"
            # Create headers
            ws.append([
                "Job Title",
                "Company", 
                "Application Date",
                "Resume Used",
                "Job Link",
                "Status"
            ])
            # Style headers
            for cell in ws[1]:
                cell.font = cell.font.copy(bold=True)
        
        # Add new row
        ws.append([
            job_title,
            company,
            application_date.strftime("%Y-%m-%d %H:%M:%S"),
            resume_path,
            job_link,
            status
        ])
        
        # Save
        wb.save(EXCEL_FILE)
        print_lg(f"📊 Application logged to {EXCEL_FILE}")
        return True
        
    except Exception as e:
        print_lg(f"❌ Failed to log application: {e}")
        return False


def get_applications_count() -> int:
    """Get the total number of logged applications."""
    if not OPENPYXL_AVAILABLE or not os.path.exists(EXCEL_FILE):
        return 0
    
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        return ws.max_row - 1  # Subtract header row
    except:
        return 0


# For standalone testing
if __name__ == "__main__":
    success = log_application(
        "Test Job Title",
        "Test Company",
        datetime.now(),
        "/path/to/resume.pdf",
        "https://linkedin.com/jobs/view/12345"
    )
    print(f"Logged: {success}")
    print(f"Total applications: {get_applications_count()}")
